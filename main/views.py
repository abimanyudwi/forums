from audioop import avg
from http.client import HTTPResponse
from multiprocessing import context
from unicodedata import category
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.sessions.middleware import SessionMiddleware
from django.http import HttpResponse
from django.utils.text import slugify
from django.urls import reverse
from hitcount.models import HitCount
from hitcount.views import HitCountMixin, HitCountDetailView
from .models import (
    Administrator,
    Author,
    Category,
    Post,
    Translasi,
    User,
    Verifikator,
    Penilaian,
    Instrumen,
    Manado,
    Perbandingan,
    Poll,
    PollResult,
    Kandidat,
    ArticleRequest,
    BeVerificator,
    TranslationNote,
)
from .utils import (
    get_link_data,
    update_views,
    compare,
    compare_db,
    kalimat_perbandingan,
    kalimat_mdo,
)
from .forms import TranslasiForms, PengumumanForms, AnnounceForms, NoteForms
from django.utils.timezone import now
from django.db.models import Avg, Max
import statistics
from statistics import mode
from el_pagination.views import AjaxListView
from el_pagination.decorators import page_template
from django import template
from users.forms import (
    NewVerifikatorForm,
    AddVerifikatorForm,
    AcceptVerifikatorForm,
    UpdateNewVerifForm,
    UpdateIndVerifForm,
    UpdateOrgVerifForm,
)
from django.utils.html import escape, strip_tags
from datetime import timedelta, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.core.mail import send_mail, BadHeaderError
from django.contrib.auth.forms import PasswordResetForm
from django.template.loader import render_to_string
from django.db.models.query_utils import Q
from django.utils.http import urlsafe_base64_encode
from django.contrib.auth.tokens import default_token_generator
from django.utils.encoding import force_bytes
from django.contrib.auth import get_user_model
from django.contrib.auth import authenticate
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError

Us = get_user_model()
# from django.http import HttpResponse
# Create your views here.
register = template.Library()


class PostCountHitDetailView(HitCountDetailView):
    model = Post
    count_hit = True


def index(request):
    userid = request.user
    post = Post.objects.all()
    num_post = Post.objects.all().count()
    num_trans = Translasi.objects.all().count()
    num_open = Post.objects.filter(status=1).count()
    num_closed = Post.objects.filter(status=0).count()
    all_post = Post.objects.values()
    all_trans = Translasi.objects.values_list("user_id", flat=True)
    # user_all = User
    # print(all_trans)
    high_user = Author.objects.get(id=mode(all_trans))
    # print(high_user)

    # uname = "Marvie"
    # polling = "Poling 1"
    # print(slugify(polling+' '+uname))

    total_translated = []

    for i in range(num_post):
        try:
            post = Post.objects.get(id=all_post[i]["id"])
            num_translated = post.translasi.all().values()
            total_translated.append(num_translated[0]["id"])
            utrans = Translasi.objects.get(id=all_trans[i])

        except:
            pass

    # print(total_translated)
    queryset = Post.objects.all().order_by("-hit_count_generic__hits")
    # queryset2 = Post.objects.all().order_by("-translasi__post_id")
    try:
        frequent = []
        freq_post = Post.translasi.through.objects.all()
        for freqp in freq_post:
            frequent.append(freqp.post_id)

        frequent = mode(frequent)
    except:
        # frequent = ""
        pass

    try:
        frequent_post = Post.objects.get(id=frequent)
    except:
        # frequent_post = queryset[0]
        frequent_post = ""

    translated = len(sorted([*set(total_translated)]))
    not_translated = num_post - translated
    try:
        last_post = Post.objects.latest("created")
    except:
        last_post = []

    most_point = Author.objects.all().order_by("-points")
    entri_manado = Manado.objects.all().count()
    entri_kalimat = Perbandingan.objects.all().count()
    user = User.objects.filter(role="USER").count()
    verifikator = User.objects.filter(role="VERIFIKATOR").count()
    try:
        user_latest = Author.objects.latest("created")
        verifikator_latest = Verifikator.objects.latest("created")
    except:
        user_latest = ""
        verifikator_latest = ""

    # print(last_post)
    context = {
        "title": "Beranda",
        "section": "Beranda",
        "post": post,
        "num_post": num_post,
        "num_trans": num_trans,
        "num_open": num_open,
        "num_closed": num_closed,
        "translated": translated,
        "not_translated": not_translated,
        "last_post": last_post,
        "frequent_post": frequent_post,
        "entri_manado": entri_manado,
        "entri_kalimat": entri_kalimat,
        "user": user,
        "userid": userid,
        "verifikator": verifikator,
        "highest": queryset[0],
        "most_point": most_point[0],
        "high_user": high_user,
        "user_latest": user_latest,
        "verifikator_latest": verifikator_latest,
    }
    return render(request, "forums/index.html", context)


@login_required
def export_entri(request):
    kata = Manado.objects.all()
    kalimat = Perbandingan.objects.all()
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response[
        "Content-Disposition"
    ] = "attachment; filename={date}-entri-manado.xlsx".format(
        date=datetime.now().strftime("%Y-%m-%d"),
    )
    workbook = Workbook()

    wordsheet = workbook.active
    wordsheet.title = "Kata"

    columns = ["Id", "Kata"]

    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = wordsheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for kt in kata:
        row_num += 1
        row = [kt.id, kt.kata]

        for col_num, cell_value in enumerate(row, 1):
            cell = wordsheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    sentencesheet = workbook.create_sheet(title="Kalimat")
    columns = [
        "Id",
        "Kalimat Indonesia",
        "Kalimat Manado",
    ]

    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = sentencesheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for kl in kalimat:
        row_num += 1
        row = [kl.id, kl.kalimat_ind, kl.kalimat_mdo]

        for col_num, cell_value in enumerate(row, 1):
            cell = sentencesheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    ws_a = wordsheet["A1"]
    ws_b = wordsheet["B1"]
    ws_a.alignment = Alignment(horizontal="center")
    ws_b.alignment = Alignment(horizontal="center")
    ws_a.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    ws_b.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    snt_a = sentencesheet["A1"]
    snt_b = sentencesheet["B1"]
    snt_c = sentencesheet["C1"]
    snt_a.alignment = Alignment(horizontal="center")
    snt_b.alignment = Alignment(horizontal="center")
    snt_c.alignment = Alignment(horizontal="center")
    snt_a.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    snt_b.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    snt_c.fill = PatternFill(
        start_color="f7f7f9",
        end_color="f7f7f9",
        fill_type="solid",
    )
    workbook.save(response)
    return response


@login_required
def profil(request):
    userid = request.user
    requested = ""

    try:
        if userid.role == "USER":
            profile = Author.objects.get(user=userid)
            author = get_object_or_404(Author, user=userid)
            requested = ArticleRequest.objects.filter(user=author).exclude(
                status="Ditolak"
            )
        elif userid.role == "VERIFIKATOR":
            profile = Verifikator.objects.get(user=userid)
        else:
            profile = Administrator.objects.get(user=userid)
    except:
        return redirect("users:update-profile")
    # user = request.user
    # del request.session["verif_status"]
    if userid.is_verified == False:
        redirect("index")
    status = (
        User.objects.filter(email=userid.email)
        .exclude(role="USER")
        .exclude(is_active=True)
        .count()
    )

    try:
        newver = BeVerificator.objects.get(user=author)
        listed = "True"
    except:
        listed = "False"
    # print(newver)

    try:
        User.objects.filter(email=userid.email).exclude(role="USER").exclude(
            is_active=True
        ).get(is_candidate=True)
        is_candidate = True
    except:
        is_candidate = False

    # print(is_candidate)
    status_user = User.objects.filter(email=userid.email).count()
    # print(status_user)

    # print(status_user)

    trans_count = Translasi.objects.filter(user_id=profile.id).count()
    ev_count = Translasi.objects.filter(penilai_id=profile.id).count()
    post_count = Post.objects.filter(user_id=userid.id).count()
    user_trans = []
    post_trans = []
    if userid.role == "USER":
        trans = Translasi.objects.filter(user=profile.id)
        for tr in trans:
            trans = tr.id
            user_trans.append(trans)
            for i in range(len(user_trans)):
                post = Translasi.objects.get(id=user_trans[i])
                post_id = post.post_set.all()
                for idp in post_id:
                    idp_str = str(idp)
                    post_trans.append(idp_str)
        trans = [*set(post_trans)]

        # print(*set(post_id))
    elif userid.role == "VERIFIKATOR":
        trans = Translasi.objects.filter(penilai_id=profile.id)
        for tr in trans:
            trans = tr.id
            user_trans.append(trans)
            for i in range(len(user_trans)):
                post = Translasi.objects.get(id=user_trans[i])
                post_id = post.post_set.all()
                for idp in post_id:
                    idp_str = str(idp)
                    post_trans.append(idp_str)
        trans = [*set(post_trans)]
    else:
        trans = Post.objects.filter(user=profile.id, tipe="Artikel")

    form = NewVerifikatorForm(request.POST, request.FILES)
    if "new-verif" in request.POST:
        print("testing")
        print(form.has_error)
        if form.is_valid():
            email = request.POST.get("email")
            proof = form.files["proof"]
            nama_lengkap = form.data["nama_lengkap"]
            field = form.data["field"]
            job_role = form.data["job_role"]
            contact = form.data["contact"]
            address = form.data["address"]
            tipe = request.POST.get("tipe")
            check_email = User.objects.filter(email=email).count()
            print(nama_lengkap)
            # if check_email >= 2:
            #     pass
            # else:
            new_ver = BeVerificator(
                nama_lengkap=nama_lengkap,
                email=email,
                proof=proof,
                field=field,
                job_role=job_role,
                contact=contact,
                address=address,
                tipe=tipe,
                user=author,
            )

            new_ver.save()
            return redirect("profil")
    if "req-berita" in request.POST:
        url = request.POST.get("request_artikel")
        now = datetime.now()
        identify = userid.username + "-" + now.strftime("%Y-%m-%d %H:%M")
        req_berita = ArticleRequest(
            url_berita=url,
            user=author,
            identifier=slugify(identify),
        )
        req_berita.save()
        return redirect("profil")
    if "cancel-form" in request.POST:
        delete = BeVerificator.objects.get(user=author.id).delete()
        return redirect("profil")

    if "on-form" in request.POST:
        if request.user.is_authenticated:
            anonym = profile
            anonym.anonym = True
            # anonym.updated = now
            anonym.save()
    if "off-form" in request.POST:
        if request.user.is_authenticated:
            anonym = profile
            anonym.anonym = False
            # anonym.updated = now
            anonym.save()

    error = "Null"
    success = "Null"
    if "reset-pass" in request.POST:
        old = request.POST.get("old")
        password = request.POST.get("password")
        confirm = request.POST.get("confirm")
        user = authenticate(request, username=request.user, password=old)
        if user is None:
            error = "Current Password Enter Mismatch!"
        else:
            try:
                if password == confirm:
                    try:
                        validate_password(password)
                        u = request.user
                        u.set_password(password)
                        u.save()
                        success = "Password has been changed!"
                        return redirect("profil")
                    except ValidationError:
                        error = "Password to Weak!"
                else:
                    error = "New Password and confirm Password mismatch!"
            except ValueError:
                error = "Password not changed!"
    # form = NewVerifikatorForm()
    # trans = post_trans
    pengumuman = Post.objects.filter(tipe="Pengumuman")

    # print(requested)
    context = {
        "title": "Profil",
        "section": "Profil",
        "userid": userid,
        "profile": profile,
        "ev_count": ev_count,
        "trans_count": trans_count,
        "post_count": post_count,
        "trans": trans,
        "status": status,
        "status_user": status_user,
        "form": form,
        "pengumuman": pengumuman,
        "is_candidate": is_candidate,
        "requested": requested,
        "listed": listed,
        "error": error,
        "success": success,
    }
    return render(request, "forums/profil.html", context)

    # return render(request, "forums/profil.html", context)


def lihat_profil_user(request, slug):
    author = get_object_or_404(Author, slug=slug)
    trans_count = Translasi.objects.filter(user_id=author.id).count()
    user_trans = []
    post_trans = []
    trans = Translasi.objects.filter(user=author.id)
    for tr in trans:
        trans = tr.id
        user_trans.append(trans)
        for i in range(len(user_trans)):
            post = Translasi.objects.get(id=user_trans[i])
            post_id = post.post_set.all()
            for idp in post_id:
                idp_str = str(idp)
                post_trans.append(idp_str)
        trans = [*set(post_trans)]

    context = {
        "title": "Profil",
        "section": "Forum",
        "author": author,
        "trans_count": trans_count,
        "trans": trans,
    }
    return render(request, "forums/lihat_profil_user.html", context)


def lihat_profil_verifikator(request, slug):
    verifikator = get_object_or_404(Verifikator, slug=slug)
    ev_count = Translasi.objects.filter(penilai_id=verifikator.id).count()
    trans = Translasi.objects.filter(penilai_id=verifikator.id)
    user_trans = []
    post_trans = []
    for tr in trans:
        trans = tr.id
        user_trans.append(trans)
        for i in range(len(user_trans)):
            post = Translasi.objects.get(id=user_trans[i])
            post_id = post.post_set.all()
            for idp in post_id:
                idp_str = str(idp)
                post_trans.append(idp_str)
        trans = [*set(post_trans)]
    context = {
        "title": "Profil",
        "section": "Forum",
        "verifikator": verifikator,
        "ev_count": ev_count,
        "trans": trans,
    }
    return render(request, "forums/lihat_profil_verifikator.html", context)


@login_required
def new_ver(request, slug):
    newv = get_object_or_404(BeVerificator, slug=slug)
    if request.user.role != "ADMIN":
        return redirect("index")

    # print(newv.status)
    # author = get_object_or_404(Author, user=newv.user)
    acceptform = AcceptVerifikatorForm(request.POST or None)
    nama = newv.nama_lengkap.lower()
    # nama_rep = nama.replace(" ", "_")
    acceptform.fields["username"].initial = slugify(nama)
    acceptform.fields["email"].initial = newv.email
    updateform = UpdateNewVerifForm(request.POST, request.FILES)
    # print(newv.nama_lengkap)
    if "reg-verif" in request.POST:
        form_1 = acceptform
        form_2 = updateform
        if form_1.is_valid() and form_2.is_valid():
            # email = request.POST.get("email")
            # print(acceptform.has_error)
            check_email = User.objects.filter(email=newv.email).count()
            if check_email >= 2:
                pass
            else:
                user = acceptform.save()
                fullname = request.POST.get("fullname")
                tipe = request.POST.get("tipe")
                field = request.POST.get("field")
                job_role = request.POST.get("job_role")
                contact = request.POST.get("contact")
                address = request.POST.get("address")
                verif = updateform.save(commit=False)
                verif.fullname = fullname
                verif.tipe = tipe
                verif.field = field
                verif.job_role = job_role
                verif.contact = contact
                verif.address = address
                verif.user = user
                verif.save()
                listed = BeVerificator.objects.get(user=newv.user)
                listed.status = "Diterima"
                listed.updated = now
                listed.save()

                subject = "Anda Diterima Menjadi Verifikator"
                email_template_name = "user-auth/accepted/accepted_email.txt"
                c = {
                    "email": acceptform.data["email"],
                    "domain": reverse("index"),
                    "site_name": "Forum Translasi",
                    "user": acceptform.data["username"],
                    "token": acceptform.data["password1"],
                    "protocol": "http",
                }
                email = render_to_string(email_template_name, c)
                try:
                    send_mail(
                        subject,
                        email,
                        "admin@example.com",
                        [newv.email],
                        fail_silently=False,
                    )
                except BadHeaderError:
                    return HttpResponse("Invalid header found.")

                return redirect("permintaan-verifikator")
    context = {
        "title": newv.nama_lengkap,
        "section": "Profil",
        "newv": newv,
        "accform": acceptform,
        "updateform": updateform,
    }
    if newv.status == "Diterima":
        return redirect("permintaan-verifikator")
    return render(request, "forums/newver.html", context)


@page_template("forums/forum_page.html")
def forum(request, template="forums/forum.html", extra_context=None):
    artikel = Post.objects.filter(tipe="Artikel").order_by("-updated")
    pengumuman = Post.objects.exclude(tipe="Artikel")
    num_posts = Post.objects.filter(tipe="Artikel").count()
    userid = request.user
    context = {
        "title": "Forum",
        "section": "Forum",
        "artikel": artikel,
        "pengumuman": pengumuman,
        "num_posts": num_posts,
        "userid": userid,
    }
    if extra_context is not None:
        context.update(extra_context)
    return render(request, template, context)


def forum_cat(request, template="forums/forum_cat.html", extra_context=None):
    kategori = Category.objects.all().order_by("urutan")
    pengumuman = Post.objects.filter(tipe="Pengumuman")
    num_cat = Category.objects.all().count()
    num_posts = Post.objects.all().count()
    userid = request.user
    context = {
        "title": "Forum",
        "section": "Forum",
        "artikel": kategori,
        "pengumuman": pengumuman,
        "num_posts": num_posts,
        "num_cat": num_cat,
        "userid": userid,
    }
    if extra_context is not None:
        context.update(extra_context)
    return render(request, template, context)


# @page_template("forums/search_page.html")
def search(request):
    return render(request, "forums/search.html")


def resend(request):
    return render(request, "forums/index.html")


def unggahan(request, slug):
    post = get_object_or_404(Post, slug=slug)
    forms = TranslasiForms(request.POST or None)
    noteforms = NoteForms(request.POST or None)
    # announceforms.fields["konten"].value = escape(strip_tags(post.konten))
    translated = ""
    translasi = ""

    try:
        author = Author.objects.get(user=request.user)
    except:
        author = ""

    if request.user.is_authenticated:
        try:
            # translated = Translasi.objects.get(user=author)
            translasi = post.translasi.get(post__pk=post.id, user=author)
            # check_author =
            # translasi = post.id
        except:
            # translated = None
            translasi = None

    try:
        if request.user.role == "VERIFIKATOR":
            verif = Verifikator.objects.get(user=request.user)
        else:
            verif = ""
    except:
        verif = ""
    # print(translated)
    # print(translasi)
    announceforms = AnnounceForms(request.POST or None)
    # announceforms.fields["konten"].widget.attrs.update(
    #     {
    #         "initValue": escape(strip_tags(post.konten)),
    #     }
    # )
    announceforms.fields["konten"].initial = post.konten
    note = ""
    try:
        note = TranslationNote.objects.get(post=post)
        noteforms.fields["content"].initial = note.content
        update = "True"
    except:
        update = "False"

    print(note)
    if request.method == "POST":
        if "perbarui-form" in request.POST:
            if announceforms.is_valid():
                content = announceforms.data["konten"]
                post.konten = content
                post.save()
                return redirect("unggahan", slug=slug)

        if "translasi-form" in request.POST:
            if forms.is_valid():
                konten = forms.data["content"]
                new_trans, created = Translasi.objects.get_or_create(
                    user=author, content=konten
                )
                post.translasi.add(new_trans.id)
                return redirect("unggahan", slug=slug)

        if "note-form" in request.POST:
            if noteforms.is_valid():
                if update == "True":
                    konten = noteforms.data["content"]
                    note.content = konten
                    note.save()
                    # post.translasi.add(new_trans.id)
                    return redirect("unggahan", slug=slug)
                elif update == "False":
                    konten = noteforms.data["content"]
                    new_note, created = TranslationNote.objects.update_or_create(
                        user=verif, content=konten, post=post
                    )
                    # post.translasi.add(new_trans.id)
                    return redirect("unggahan", slug=slug)

        # if "translasi-form" in request.POST:
        #     if forms.is_valid():
        #         konten = forms.data["content"]
        #         new_trans, created = Translasi.objects.get_or_create(
        #             user=author, content=konten
        #         )
        #         post.translasi.add(new_trans.id)
        #         return redirect("unggahan", slug=slug)

        if "vote-form" in request.POST:
            poll_slug = request.POST.get("polls-slug")
            poll = get_object_or_404(Poll, slug=poll_slug)
            pick = request.POST.get("kandidat")
            kandidat = get_object_or_404(Kandidat, slug=pick)
            voter = request.POST.get("polls-user")
            voter_uname = User.objects.get(id=voter)
            identify = voter_uname.username + " " + poll.nama_polling
            new_vote, created = PollResult.objects.get_or_create(
                identifier=slugify(identify),
                user_id=voter,
                polls=pick,
            )
            # poll.translasi.add(new_trans.id)
            kandidat.result.add(new_vote.id)

            return redirect("unggahan", slug=slug)

    if "point-form" in request.POST:
        if request.user.is_authenticated:
            pt1 = request.POST.get("pt1")
            pt2 = request.POST.get("pt2")
            pt3 = request.POST.get("pt3")
            pt4 = request.POST.get("pt4")
            points = (
                (float(pt1) * 0.4)
                + (float(pt2) * 0.3)
                + (float(pt3) * 0.2)
                + (float(pt4) * 0.1)
            )
            points = round(points, 3)
            tr_id = request.POST.get("trans-id")
            tr_obj = Translasi.objects.get(id=tr_id)
            author = Verifikator.objects.get(user=request.user)
            tr_obj.poin = points
            tr_obj.penilai = author
            tr_obj.updated = now
            tr_obj.save()
            user_trans = (
                Translasi.objects.filter(user_id=tr_obj.user)
                .exclude(poin=0)
                .aggregate(Avg("poin"))
            )
            points_avg = user_trans

            user_obj = Author.objects.get(id=tr_obj.user.id)
            user_obj.points = round(points_avg["poin__avg"], 2)
            user_obj.save()
            return redirect("unggahan", slug=post.slug)

    if "close-form" in request.POST:
        if request.user.is_authenticated:
            post = Post.objects.get(id=post.id)
            post.status = False
            post.updated = now
            post.dinilai = False
            post.save()

    if "finish-form" in request.POST:
        if request.user.is_authenticated:
            post = Post.objects.get(id=post.id)
            post.status = False
            post.updated = now
            post.dinilai = True
            post.save()
            # post = post.id
            trans = Post.objects.get(id=post.id)
            tr = trans.translasi.all().aggregate(Max("poin"))
            poin = tr
            pts = poin["poin__max"]
            record = trans.translasi.filter(poin=pts).values("id")
            # record = record.latest('created')
            # tr = tr.id
            tr_id = record[0]["id"]
            translasi = Translasi.objects.get(id=tr_id)
            translasi.best = True
            translasi.save()
            # Algoritma komparasi kata
            translasi = translasi.content
            hasil_banding = compare(translasi)
            hasil = hasil_banding["compare"]
            kata_db = Manado.objects.all().distinct()
            list_kata = []
            for kata_db in kata_db:
                kata_db = str(kata_db)
                list_kata.append(kata_db)
            banding_kata = compare_db(hasil, list_kata)
            banding_kata = banding_kata["compare"]
            # print(banding_kata)
            if banding_kata == None:
                pass
            else:
                for hsl in banding_kata:
                    Manado.objects.create(kata=hsl)
            kalimat = kalimat_perbandingan(post.konten, translasi)
            kalimat_ind = kalimat["ind"]
            kalimat_mdo = kalimat["mdo"]
            zipped = zip(kalimat_ind, kalimat_mdo)
            for ind, mdo in zipped:
                Perbandingan.objects.get_or_create(kalimat_ind=ind, kalimat_mdo=mdo)

    penilaian = {
        "penilaian1": Penilaian.objects.get(id=1),
        "penilaian2": Penilaian.objects.get(id=2),
        "penilaian3": Penilaian.objects.get(id=3),
        "penilaian4": Penilaian.objects.get(id=4),
    }

    instrumen1 = (
        Instrumen.objects.filter(penilaian__id=penilaian["penilaian1"].id).values(),
    )
    instrumen2 = (
        Instrumen.objects.filter(penilaian__id=penilaian["penilaian2"].id).values(),
    )
    instrumen3 = (
        Instrumen.objects.filter(penilaian__id=penilaian["penilaian3"].id).values(),
    )
    instrumen4 = (
        Instrumen.objects.filter(penilaian__id=penilaian["penilaian4"].id).values(),
    )

    dinilai = post.translasi.exclude(poin=0.0).count()
    # print(dinilai)
    # for key, value in inst.items:
    # print(num_instrumen1[0])
    try:
        poll_active = Poll.objects.get(is_active=1)
        poll_status = "exists"
    except:
        poll_active = Poll.objects.filter(is_active=0).latest("updated")
        poll_status = "not existed"

    # kandidat = poll_active.kandidat.all()
    permintaan = (
        User.objects.filter(is_active=0).exclude(role="USER").exclude(is_candidate=1)
    )
    # add-candidate-form
    if "add-candidate-form" in request.POST:
        if request.user.is_authenticated:
            if request.user.role == "ADMIN":
                email = request.POST.get("email")
                get_objects = User.objects.filter(email=email).exclude(
                    role="Verifikator"
                )
                # print(get_objects)
                for un in get_objects:
                    name = Author.objects.get(user=un)
                    # print(name.user.username)
                uname = name.user.username
                pollname = request.POST.get("pollname")
                # email = request.POST.get("email")
                # print(uname)
                uid = User.objects.get(username=uname)
                author_id = Author.objects.get(user=uid.id)
                # print(uid.id)
                slug = slugify(pollname + " " + str(uname))
                instance = (
                    User.objects.filter(email=email).exclude(role="USER").values()
                )
                new_ins = User.objects.get(id=instance[0]["id"])
                # print(instance[0]['id'])
                new_ins.is_candidate = True
                new_ins.save()
                new_candidate, created = Kandidat.objects.get_or_create(
                    user=author_id, slug=slug, is_approved=0
                )
                poll_active.kandidat.add(new_candidate.id)

                return redirect("unggahan", slug=post.slug)

    try:
        leader = PollResult.objects.filter(polls__icontains=poll_active.slug)
        top_vote = []
        for lead in leader:
            top = lead.polls
            top_vote.append(top)
        winner = Kandidat.objects.get(slug=mode(top_vote))
        total_vote = leader.count()
    except:
        winner = "None"
        total_vote = "None"

    if "close-poll" in request.POST:
        if request.user.is_authenticated:
            slug_poll = request.POST.get("poll-slug")
            poll = Poll.objects.get(slug=slug_poll)
            # print(poll)
            poll.is_active = False
            poll.save()
            winning_author = Author.objects.get(id=winner.user.id)
            # print(winning_author)
            winning_user = User.objects.filter(role="VERIFIKATOR").get(
                email=winning_author.user.email
            )
            winning_user.is_active = True
            winning_user.save()
            cand_close = User.objects.filter(is_candidate=True)
            for cand in cand_close:
                cand.is_candidate = False
                cand.save()
            winning_candidate = Kandidat.objects.get(slug=winner.slug)
            winning_candidate.is_approved = True
            winning_candidate.save()

            return redirect("unggahan", slug=slug)
    if "start-poll" in request.POST:
        try:
            Poll.objects.get(is_active=True)
            return redirect("unggahan", slug=slug)

        except:
            pass
        nama_polling = request.POST.get("poll-name")
        creator = request.POST.get("creator")
        poll_ke = request.POST.get("poll-num")
        new_polls, created = Poll.objects.get_or_create(
            nama_polling=nama_polling,
            creator_id=creator,
            poll_ke=poll_ke,
        )
        # poll.translasi.add(new_trans.id)

        return redirect("unggahan", slug=slug)

    num_user = User.objects.filter(role="USER").count()
    # print(modes)

    identifier = slugify(request.user.username + " " + poll_active.nama_polling)
    try:
        PollResult.objects.get(identifier=identifier)
        is_voted = True
    except:
        is_voted = False

    try:
        vresult = PollResult.objects.filter(polls__icontains=poll_active.slug).get(
            user=request.user
        )
        votes = Kandidat.objects.get(result=vresult)
    except:
        votes = None

    next_poll = poll_active.poll_ke + 1
    # print(next_poll)
    # print(is_voted)
    candidate = poll_active.kandidat.all()
    context = {
        "title": "Unggahan",
        "section": "Forum",
        "post": post,
        "forms": forms,
        "announceforms": announceforms,
        "translated": translated,
        "translasi": translasi,
        "author": author,
        "penilaian": penilaian,
        "instrumen1": instrumen1,
        "instrumen2": instrumen2,
        "instrumen3": instrumen3,
        "instrumen4": instrumen4,
        "userid": request.user,
        "dinilai": dinilai,
        "poll_active": poll_active,
        "permintaan": permintaan,
        "is_voted": is_voted,
        "candidate": candidate,
        "votes": votes,
        "poll_status": poll_status,
        "next_poll": next_poll,
        "winner": winner,
        "total_vote": total_vote,
        "num_user": num_user,
        "noteforms": noteforms,
    }
    update_views(request, post)
    return render(request, "forums/unggahan.html", context)


@login_required
def unggah_artikel(request):
    user = request.user
    role = user.role
    admin = "ADMIN"
    if role != admin:
        return redirect("users:logout")

    error = ""
    try:
        if request.method == "POST":
            url = request.POST.get("url")
            artikel = get_link_data(url)
            request.session["article_url"] = artikel
            return redirect("konten-artikel")
    except AttributeError:
        error = "Tidak dapat mengambil artikel"
    except:
        error = "Error"

    context = {"title": "Unggah Artikel", "section": "Forum", "error": error}
    return render(request, "forums/unggah_artikel.html", context)


@login_required
def konten_artikel(request):
    category = Category.objects.all()
    identifier = ""

    artikel = request.session.get("article_url")
    try:
        identifier = request.session.get("req_iden")
    except:
        identifier = ""

    # print(identifier)
    if artikel is None:
        return redirect("unggah-artikel")

    if request.method == "POST":
        judul = request.POST.get("judul")
        portal = request.POST.get("portal")
        url = request.POST.get("url")
        konten = request.POST.get("konten-area")
        get_kategori = request.POST.get("kategori")
        kategori = get_object_or_404(Category, id=get_kategori)
        author = Administrator.objects.get(user=request.user)
        user = author
        post = Post(
            judul=judul,
            portal=portal,
            url=url,
            konten=konten,
            user=user,
            kategori=kategori,
        )
        post.save()
        try:
            req_art = ArticleRequest.objects.get(identifier=identifier)
            req_art.status = "Diterima"
            req_art.updated = now
            req_art.save()
            del request.session["req_iden"]
            del request.session["article_url"]
            return redirect("requested-article")
        except:
            del request.session["article_url"]
            return redirect("forum")
        # kategori = post_category.objects.create(category_id=kategori)
        # post.kategori.add(kategori)
        del request.session["article_url"]
        return redirect("forum")
    context = {
        "title": "Konten Artikel",
        "section": "Forum",
        "category": category,
        "artikel": artikel,
        "identifier": identifier
        # "error":error,
    }
    return render(request, "forums/generated_article.html", context)


@login_required
def konten_pengumuman(request):
    forms = PengumumanForms(request.POST or None)
    if request.method == "POST":
        judul = request.POST.get("judul")
        portal = "None"
        url = "None"
        konten = forms.data["konten"]
        kategori = get_object_or_404(Category, id="8")
        tipe = "Pengumuman"
        author = Administrator.objects.get(user=request.user)
        post = Post(
            judul=judul,
            portal=portal,
            url=url,
            konten=konten,
            user=author,
            tipe=tipe,
            kategori=kategori,
        )
        post.save()
        return redirect("forum")
    context = {
        "title": "Konten Artikel",
        "section": "Forum",
        "forms": forms,
        # "error":error,
    }
    return render(request, "forums/tambah_pengumuman.html", context)


@login_required
def daftar_user(request):
    # post = Post.objects.get(id="2")
    # trans = Translasi.objects.get(id="12")
    # tr = trans.content.strip()
    # kalimat = kalimat_perbandingan(post.konten.strip(), trans.content.strip())
    # print(kalimat['mdo'])

    if request.user.role != "ADMIN":
        return redirect("forum")

    accform = AddVerifikatorForm(request.POST or None)
    indform = UpdateIndVerifForm(request.POST, request.FILES)
    orgform = UpdateOrgVerifForm(request.POST, request.FILES)

    if "reg-ind" in request.POST:
        form_1 = accform
        form_2 = indform
        if form_1.is_valid() and form_2.is_valid():
            print(form_1.has_error)
            print(form_2.has_error)
            tipe = request.POST.get("tipe")
            user = accform.save(commit=False)
            user.role = "VERIFIKATOR"
            user.save()
            indver = indform.save(commit=False)
            indver.tipe = tipe
            indver.user = user
            indver.save()
            return redirect("daftar-user")
    if "reg-org" in request.POST:
        form_1 = accform
        form_2 = orgform
        if form_1.is_valid() and form_2.is_valid():
            tipe = request.POST.get("tipe")
            user = accform.save(commit=False)
            user.role = "VERIFIKATOR"
            user.save()
            orgver = orgform.save(commit=False)
            orgver.tipe = tipe
            orgver.user = user
            orgver.save()
            return redirect("daftar-user")
    author = Author.objects.all()
    verifikator = User.objects.filter(role="VERIFIKATOR").exclude(
        username=("test-ver,org-ver")
    )
    context = {
        "title": "Daftar User",
        "section": "Profil",
        "verifikator": verifikator,
        "num_verifikator": verifikator.count(),
        "author": author,
        "num_author": author.count(),
        "accform": accform,
        "indform": indform,
        "orgform": orgform,
    }
    return render(request, "forums/daftar_user.html", context)


@login_required
def requested_article(request):
    # post = Post.objects.get(id="2")
    # trans = Translasi.objects.get(id="12")
    # tr = trans.content.strip()
    # kalimat = kalimat_perbandingan(post.konten.strip(), trans.content.strip())
    # print(kalimat['mdo'])

    if request.user.role != "ADMIN":
        return redirect("forum")

    if "generate-article" in request.POST:
        try:
            if request.method == "POST":
                identifier = request.POST.get("generate")
                req = get_object_or_404(ArticleRequest, identifier=identifier)
                artikel = get_link_data(req.url_berita)
                request.session["article_url"] = artikel
                request.session["req_iden"] = identifier
                return redirect("konten-artikel")
        except AttributeError:
            error = "Tidak dapat mengambil artikel"
        except:
            error = "Error"

    req_article = ArticleRequest.objects.all().exclude(status="Ditolak")
    context = {
        "title": "Requested Article",
        "section": "Profil",
        "req_article": req_article,
        "num_requested": req_article.count(),
        # "form": form,
    }
    return render(request, "forums/reqed_article.html", context)


@login_required
def permintaan_verifikator(request):
    # post = Post.objects.get(id="2")
    # trans = Translasi.objects.get(id="12")
    # tr = trans.content.strip()
    # kalimat = kalimat_perbandingan(post.konten.strip(), trans.content.strip())
    # print(kalimat['mdo'])

    if request.user.role != "ADMIN":
        return redirect("forum")

    user = BeVerificator.objects.all()

    # print(existed)
    context = {
        "title": "Daftar User",
        "section": "Profil",
        "user": user,
        "num_user": user.count(),
    }
    return render(request, "forums/permintaan_verifikator.html", context)


@login_required
def verifikasi_terjemahan(request):
    # post = Post.objects.get(id="2")
    # trans = Translasi.objects.get(id="12")
    # tr = trans.content.strip()
    # kalimat = kalimat_perbandingan(post.konten.strip(), trans.content.strip())
    # print(kalimat['mdo'])

    if request.user.role != "VERIFIKATOR":
        return redirect("forum")

    post = (
        Post.objects.filter(tipe="Artikel", dinilai=False)
        .exclude(status=1)
        .order_by("-updated")
    )

    # print(existed)
    context = {
        "title": "Daftar User",
        "section": "Profil",
        "post": post,
        "num_post": post.count(),
    }
    return render(request, "forums/verifikasi_terjemahan.html", context)


@login_required
def daftar_kata(request):
    if request.user.role != "ADMIN":
        return redirect("forum")

    kata = Manado.objects.all()
    # verifikator = Verifikator.objects.all()
    context = {
        "title": "Daftar User",
        "section": "Profil",
        "kata": kata,
        "num_kata": kata.count(),
    }
    return render(request, "forums/daftar_kata.html", context)


def password_reset_request(request):
    if request.method == "POST":
        password_reset_form = PasswordResetForm(request.POST)
        if password_reset_form.is_valid():
            data = password_reset_form.cleaned_data["email"]
            uname = password_reset_form.cleaned_data["username"]
            associated_users = Us.objects.filter(Q(email=data) & Q(username=uname))
            if associated_users.exists():
                for user in associated_users:
                    subject = "Password Reset Requested"
                    email_template_name = "user-auth/password/password_reset_email.txt"
                    c = {
                        "email": user.email,
                        "domain": "127.0.0.1:8000",
                        "site_name": "Website",
                        "uid": urlsafe_base64_encode(force_bytes(user.pk)),
                        "user": user,
                        "token": default_token_generator.make_token(user),
                        "protocol": "http",
                    }
                    email = render_to_string(email_template_name, c)
                    try:
                        send_mail(
                            subject,
                            email,
                            "admin@example.com",
                            [user.email],
                            fail_silently=False,
                        )
                    except BadHeaderError:
                        return HttpResponse("Invalid header found.")
                    return redirect("/password_reset/done/")
    password_reset_form = PasswordResetForm()
    return render(
        request=request,
        template_name="user-auth/password/password_reset.html",
        context={"password_reset_form": password_reset_form},
    )


# @login_required
# def kategori(request, slug):
#     kategori = get_object_or_404(Category, slug=slug)
#     user = request.user
#     post = Post.objects.all().filter(kategori=kategori)
#     context = {
#         "title": kategori,
#         "section": kategori,
#         "post": post,
#         "user": user.role,
#     }
#     return render(request, "manager/berkas-page.html", context)


def kategori(request, slug):
    kategori = get_object_or_404(Category, slug=slug)
    artikel = Post.objects.filter(tipe="Artikel", kategori=kategori).order_by(
        "-updated"
    )
    num_posts = Post.objects.filter(tipe="Artikel", kategori=kategori).count()
    userid = request.user
    context = {
        "title": "Forum",
        "section": "Forum",
        "artikel": artikel,
        "num_posts": num_posts,
        "userid": userid,
        "kategori": kategori,
    }

    return render(request, "forums/kategori.html", context)
